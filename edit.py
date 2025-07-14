import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import time

def color_rows_based_on_column_v(file_path, output_path, words_to_check):
    
    # Define colors for each word (Red, Green, Blue)
    colors = [
        PatternFill(start_color='f2564b', end_color='f2564b', fill_type='solid'),  # Red
        PatternFill(start_color='f7f723', end_color='f7f723', fill_type='solid'),  # Yellow
        PatternFill(start_color='07f7f7', end_color='07f7f7', fill_type='solid')   # Blue
    ]
    
    # Color names for messages
    color_names = ["RED", "YELLOW", "BLUE"]
    
    print(f"\nStarting Excel file processing: {file_path}")
    print(f"Looking for words: {', '.join(words_to_check)} in column V")
    print("-" * 60)
    
    # Load the workbook
    try:
        start_time = time.time()
        wb = openpyxl.load_workbook(file_path)
        load_time = time.time() - start_time
        print(f"File loaded successfully in {load_time:.2f} seconds")
    except FileNotFoundError:
        print(f"❌ Error: File '{file_path}' not found.")
        return
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return
    
    # Process each worksheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nProcessing worksheet: '{sheet_name}'")
        print(f"Total rows to process: {ws.max_row}")
        print("-" * 40)
        
        # Find column V (22nd column)
        col_v = get_column_letter(22)
        colored_rows = 0
        
        # Iterate through rows in column V
        for row in range(1, ws.max_row + 1):
            cell = ws[f"{col_v}{row}"]
            cell_value = str(cell.value).strip().lower() if cell.value else ""
            
            # Default message (no match)
            message = f"Row {row}: No matching words found"
            colored = False
            
            # Check for each word and color the row accordingly
            for i, word in enumerate(words_to_check):
                if word.lower() in cell_value:
                    # Color the entire row
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = colors[i]
                    message = f"Row {row}: Found '{word}' - colored {color_names[i]}"
                    colored = True
                    colored_rows += 1
                    break  # Stop checking other words if found
            
            # Print message for current row
            print(message)
            time.sleep(0.05)  # Small delay to make messages readable
            
        # Worksheet summary
        print("-" * 40)
        print(f"Worksheet '{sheet_name}' processed:")
        print(f"Total rows: {ws.max_row}")
        print(f"Colored rows: {colored_rows}")
        print(f"Percentage colored: {(colored_rows/ws.max_row)*100:.1f}%")
    
    # Save the modified workbook
    print("\nSaving results...")
    try:
        save_start = time.time()
        wb.save(output_path)
        save_time = time.time() - save_start
        print(f"✅ File saved successfully to: {output_path}")
        print(f"Save operation took {save_time:.2f} seconds")
        total_time = time.time() - start_time
        print(f"\nTotal processing time: {total_time:.2f} seconds")
    except PermissionError:
        print("❌ Error: Permission denied. Please close the Excel file if it's open.")
    except Exception as e:
        print(f"❌ Error saving file: {e}")

